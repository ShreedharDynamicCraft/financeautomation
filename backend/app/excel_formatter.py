import pandas as pd
import logging
import io
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, NamedStyle
)
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter
from typing import Dict, Any
import os
from datetime import datetime

logger = logging.getLogger(__name__)

def create_formatted_excel(extracted_data: dict, template_type: str, output_path: str) -> str:
    """
    Create a beautifully formatted, colorful, and professional Excel report
    
    Args:
        extracted_data: Dictionary containing extracted data
        template_type: Template type used ("Extraction Template 1" or "Extraction Template 2")
        output_path: Path where Excel file should be saved
        
    Returns:
        Path to created Excel file
    """
    buffer = create_formatted_excel_buffer(extracted_data, template_type)
    
    # Save buffer to file
    with open(output_path, 'wb') as f:
        f.write(buffer.getvalue())
    
    logger.info(f"Beautifully formatted Excel file saved to: {output_path}")
    return output_path

def create_formatted_excel_buffer(extracted_data: dict, template_type: str) -> io.BytesIO:
    """
    Create a beautifully formatted, colorful, and professional Excel report as BytesIO buffer
    
    Args:
        extracted_data: Dictionary containing extracted data
        template_type: Template type used ("Extraction Template 1" or "Extraction Template 2")
        
    Returns:
        BytesIO buffer containing Excel file
    """
    try:
        logger.info(f"Creating beautifully formatted Excel report for {template_type}")
        
        # Create a new workbook
        wb = Workbook()
        
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # Define professional color scheme
        colors = {
            'primary': '1f4e79',      # Deep Blue
            'secondary': '70ad47',     # Green
            'accent': 'ffc000',        # Gold
            'light_blue': 'd6dcf0',    # Light Blue
            'light_green': 'e2efda',   # Light Green
            'light_gold': 'fff2cc',    # Light Gold
            'white': 'ffffff',         # White
            'dark_text': '2c3e50',     # Dark Gray
            'header_text': 'ffffff'    # White for headers
        }
        
        # Create custom styles
        setup_custom_styles(wb, colors)
        
        # Create overview/summary sheet first
        create_overview_sheet(wb, extracted_data, template_type, colors)
        
        # Process each sheet in the extracted data
        sheet_order = get_sheet_order(template_type)
        
        for sheet_name in sheet_order:
            if sheet_name in extracted_data and extracted_data[sheet_name]:
                try:
                    create_formatted_sheet(
                        wb, sheet_name, extracted_data[sheet_name], colors, template_type
                    )
                    logger.info(f"Created beautifully formatted sheet: {sheet_name}")
                except Exception as e:
                    logger.error(f"Error creating sheet {sheet_name}: {e}")
                    continue
        
        # Add any remaining sheets not in the predefined order
        for sheet_name, sheet_data in extracted_data.items():
            if sheet_name not in sheet_order and sheet_data:
                try:
                    create_formatted_sheet(
                        wb, sheet_name, sheet_data, colors, template_type
                    )
                    logger.info(f"Created additional sheet: {sheet_name}")
                except Exception as e:
                    logger.error(f"Error creating additional sheet {sheet_name}: {e}")
                    continue
        
        # Save to BytesIO buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        logger.info(f"Beautifully formatted Excel buffer created successfully")
        return buffer
        
    except Exception as e:
        logger.error(f"Excel formatting failed: {e}")
        raise

def setup_custom_styles(wb: Workbook, colors: Dict[str, str]):
    """Setup custom named styles for consistent formatting"""
    
    # Header style - Professional blue header
    header_style = NamedStyle(name="header_style")
    header_style.font = Font(
        name='Calibri', size=12, bold=True, color=colors['header_text']
    )
    header_style.fill = PatternFill(
        start_color=colors['primary'], end_color=colors['primary'], fill_type='solid'
    )
    header_style.alignment = Alignment(
        horizontal='center', vertical='center', wrap_text=True
    )
    header_style.border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='medium', color='000000')
    )
    wb.add_named_style(header_style)
    
    # Data style - Clean data presentation
    data_style = NamedStyle(name="data_style")
    data_style.font = Font(name='Calibri', size=11, color=colors['dark_text'])
    data_style.alignment = Alignment(horizontal='left', vertical='center')
    data_style.border = Border(
        left=Side(style='thin', color='cccccc'),
        right=Side(style='thin', color='cccccc'),
        top=Side(style='thin', color='cccccc'),
        bottom=Side(style='thin', color='cccccc')
    )
    wb.add_named_style(data_style)
    
    # Number style - For monetary values
    number_style = NamedStyle(name="number_style")
    number_style.font = Font(name='Calibri', size=11, color=colors['dark_text'])
    number_style.alignment = Alignment(horizontal='right', vertical='center')
    number_style.number_format = '#,##0'
    number_style.border = Border(
        left=Side(style='thin', color='cccccc'),
        right=Side(style='thin', color='cccccc'),
        top=Side(style='thin', color='cccccc'),
        bottom=Side(style='thin', color='cccccc')
    )
    wb.add_named_style(number_style)
    
    # Currency style - For monetary values with currency symbol
    currency_style = NamedStyle(name="currency_style")
    currency_style.font = Font(name='Calibri', size=11, color=colors['dark_text'])
    currency_style.alignment = Alignment(horizontal='right', vertical='center')
    currency_style.number_format = '$#,##0'
    currency_style.border = Border(
        left=Side(style='thin', color='cccccc'),
        right=Side(style='thin', color='cccccc'),
        top=Side(style='thin', color='cccccc'),
        bottom=Side(style='thin', color='cccccc')
    )
    wb.add_named_style(currency_style)

def create_overview_sheet(wb: Workbook, extracted_data: dict, template_type: str, colors: Dict[str, str]):
    """Create a beautiful overview/dashboard sheet"""
    
    overview_ws = wb.create_sheet("📊 Executive Summary", 0)
    
    # Add title
    overview_ws['B2'] = f"PDF Extraction Report - {template_type}"
    overview_ws['B2'].font = Font(name='Calibri', size=18, bold=True, color=colors['primary'])
    
    # Add generation info
    overview_ws['B4'] = f"Generated on: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}"
    overview_ws['B4'].font = Font(name='Calibri', size=11, color=colors['dark_text'])
    
    overview_ws['B5'] = f"Template Type: {template_type}"
    overview_ws['B5'].font = Font(name='Calibri', size=11, color=colors['dark_text'])
    
    # Create summary table
    summary_data = []
    for sheet_name, sheet_data in extracted_data.items():
        if isinstance(sheet_data, list):
            record_count = len(sheet_data)
            summary_data.append([sheet_name, record_count, "✓ Processed"])
        else:
            summary_data.append([sheet_name, "N/A", "⚠ Empty"])
    
    # Headers for summary table
    overview_ws['B8'] = "Sheet Name"
    overview_ws['C8'] = "Records"
    overview_ws['D8'] = "Status"
    
    # Apply header styling
    for col in ['B8', 'C8', 'D8']:
        cell = overview_ws[col]
        cell.style = "header_style"
    
    # Add summary data
    row = 9
    for sheet_name, record_count, status in summary_data:
        overview_ws[f'B{row}'] = sheet_name
        overview_ws[f'C{row}'] = record_count
        overview_ws[f'D{row}'] = status
        
        # Apply data styling
        for col in ['B', 'C', 'D']:
            cell = overview_ws[f'{col}{row}']
            cell.style = "data_style"
            
        # Color code status
        status_cell = overview_ws[f'D{row}']
        if "✓" in status:
            status_cell.fill = PatternFill(
                start_color=colors['light_green'], 
                end_color=colors['light_green'], 
                fill_type='solid'
            )
        elif "⚠" in status:
            status_cell.fill = PatternFill(
                start_color=colors['light_gold'], 
                end_color=colors['light_gold'], 
                fill_type='solid'
            )
        
        row += 1
    
    # Auto-fit columns
    for column in overview_ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 3, 50)
        overview_ws.column_dimensions[column_letter].width = adjusted_width

def create_formatted_sheet(wb: Workbook, sheet_name: str, sheet_data: list, colors: Dict[str, str], template_type: str):
    """Create a beautifully formatted sheet with professional styling"""
    
    if not sheet_data:
        return
    
    # Clean sheet name for Excel compatibility
    clean_name = clean_sheet_name(sheet_name)
    ws = wb.create_sheet(clean_name)
    
    # Convert to DataFrame for easier manipulation
    df = pd.DataFrame(sheet_data)
    
    if df.empty:
        return
    
    # Add title row
    ws['A1'] = sheet_name
    ws['A1'].font = Font(name='Calibri', size=14, bold=True, color=colors['primary'])
    
    # Add headers starting from row 3
    for col_idx, column_name in enumerate(df.columns, 1):
        cell = ws.cell(row=3, column=col_idx, value=column_name)
        cell.style = "header_style"
    
    # Add data starting from row 4
    for row_idx, (_, row_data) in enumerate(df.iterrows(), 4):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            
            # Apply appropriate styling based on data type
            if isinstance(value, (int, float)) and value != 0:
                if is_currency_field(df.columns[col_idx-1]):
                    cell.style = "currency_style"
                else:
                    cell.style = "number_style"
            else:
                cell.style = "data_style"
    
    # Auto-fit columns
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 3, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Add alternating row colors for better readability
    add_alternating_row_colors(ws, len(df) + 3, len(df.columns), colors)
    
    # Add conditional formatting for numerical columns
    add_conditional_formatting(ws, df, colors)

def clean_sheet_name(name: str) -> str:
    """Clean sheet name to comply with Excel naming rules"""
    # Excel sheet names cannot contain: \/:*?"<>|
    # and must be 31 characters or less
    invalid_chars = ['\\', '/', ':', '*', '?', '"', '<', '>', '|']
    clean_name = name
    for char in invalid_chars:
        clean_name = clean_name.replace(char, '_')
    return clean_name[:31]

def is_currency_field(column_name: str) -> bool:
    """Determine if a field should be formatted as currency"""
    currency_keywords = [
        'value', 'amount', 'capital', 'investment', 'revenue', 
        'ebitda', 'income', 'expenses', 'nav', 'commitment',
        'cost', 'price', 'fee', 'debt', 'cash', 'aum'
    ]
    return any(keyword in column_name.lower() for keyword in currency_keywords)

def add_alternating_row_colors(ws, last_row: int, last_col: int, colors: Dict[str, str]):
    """Add alternating row colors for better readability"""
    light_fill = PatternFill(
        start_color=colors['light_blue'], 
        end_color=colors['light_blue'], 
        fill_type='solid'
    )
    
    for row in range(4, last_row + 1):  # Start from data rows
        if row % 2 == 0:  # Even rows
            for col in range(1, last_col + 1):
                cell = ws.cell(row=row, column=col)
                if cell.fill.start_color.index == '00000000':  # No existing fill
                    cell.fill = light_fill

def add_conditional_formatting(ws, df: pd.DataFrame, colors: Dict[str, str]):
    """Add conditional formatting to numerical columns"""
    for col_idx, column_name in enumerate(df.columns, 1):
        # Check if column contains numerical data
        if df[column_name].dtype in ['int64', 'float64']:
            column_letter = get_column_letter(col_idx)
            data_range = f"{column_letter}4:{column_letter}{len(df) + 3}"
            
            # Create a color scale rule (green to yellow to red)
            color_scale_rule = ColorScaleRule(
                start_type='min',
                start_color=colors['secondary'],  # Green
                mid_type='percentile',
                mid_value=50,
                mid_color=colors['accent'],       # Gold
                end_type='max',
                end_color='e74c3c'               # Red
            )
            ws.conditional_formatting.add(data_range, color_scale_rule)

def get_sheet_order(template_type: str) -> list:
    """Define the order of sheets for professional presentation"""
    
    if template_type == "Extraction Template 1":
        return [
            "Fund Data",
            "Fund Manager", 
            "Fund Financial Position",
            "Company Investment Positions",
            "Initial Investments",
            "Company Valuation",
            "Company Financials",
            "Fund Companies",
            "LP cashflows",
            "Doc Summary"
        ]
    elif template_type == "Extraction Template 2":
        return [
            "Portfolio Summary",
            "Schedule of Investments",
            "Portfolio Company profile",
            "Portfolio Company Financials",
            "Statement of Operations",
            "Statement of Cashflows", 
            "PCAP Statement",
            "Footnotes",
            "Doc Summary"
        ]
    else:
        return []